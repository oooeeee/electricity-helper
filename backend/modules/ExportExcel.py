import itertools
from openpyxl import Workbook
from openpyxl.styles import Alignment
from tempfile import NamedTemporaryFile
from .common import parse_date, sort_houses

DATA_COUNT_TO_EXPORT = 4


def centered(cell):
    cell.alignment = Alignment(horizontal="center", vertical="center")


class Rows:
    street_name = 1
    dates = 2
    price_types = 3
    houses_start = 4


class ExportExcel:
    def __init__(self, storage, price_types):
        self.storage = storage
        self.price_types = price_types

    def _get_known_dates(self, street_name):
        dates = set()
        for house_info in self.storage.get_street(street_name, last_data_count=DATA_COUNT_TO_EXPORT * 2).values():
            dates.update(date_info['date'] for date_info in house_info)
        return sorted(dates, key=parse_date)[-DATA_COUNT_TO_EXPORT:]

    @staticmethod
    def _get_sheet_for_street(book, street_name):
        sheet = book.create_sheet(title=street_name)
        sheet.merge_cells(start_row=Rows.street_name, start_column=1, end_row=Rows.street_name, end_column=(DATA_COUNT_TO_EXPORT * 2) + 1)
        centered(sheet.cell(column=1, row=Rows.street_name, value=street_name))
        return sheet

    def _add_row_dates(self, sheet, street_name):
        known_dates = sorted(self._get_known_dates(street_name), key=parse_date)
        sheet.cell(column=1, row=Rows.dates, value="#")
        for column, date in enumerate(known_dates, start=0):
            column = (column * len(self.price_types) + 2)
            sheet.merge_cells(start_column=column, start_row=Rows.dates, end_column=column + 1, end_row=Rows.dates)
            centered(sheet.cell(column=column, row=Rows.dates, value=date))
        return known_dates

    def _add_row_price_types(self, sheet, dates_count):
        column = itertools.count(2)
        for _ in range(dates_count):
            for price_type in self.price_types:
                sheet.cell(column=next(column), row=Rows.price_types, value=price_type)

    def _add_row_house(self, sheet, row, dates, house, house_dates):
        column = itertools.count(1)
        sheet.cell(column=next(column), row=row, value=house)
        house_dates = {date_info['date']: date_info for date_info in house_dates}
        for date in dates:
            for price_type in self.price_types:
                value = house_dates.get(date, {}).get(price_type, '')
                sheet.cell(column=next(column), row=row, value=value)

    @staticmethod
    def _remove_unknown_sheetnames(book, allowed_sheet_names):
        for sheet in book.worksheets:
            if sheet.title not in allowed_sheet_names:
                book.remove(sheet)

    @staticmethod
    def _get_workbook_source(book):
        with NamedTemporaryFile() as tmp:
            book.save(tmp.name)
            tmp.seek(0)
            return tmp.read()

    def export_to_excel(self):
        book = Workbook()
        street_names = self.storage.get_street_names()
        for street_name in street_names:
            sheet = self._get_sheet_for_street(book, street_name)
            dates = self._add_row_dates(sheet, street_name)
            houses = self.storage.get_street(street_name, last_data_count=DATA_COUNT_TO_EXPORT)
            self._add_row_price_types(sheet, len(dates))
            for row_index, house in enumerate(sort_houses(houses.keys()), start=Rows.houses_start):
                self._add_row_house(sheet, row_index, dates, house, houses[house])
        self._remove_unknown_sheetnames(book, street_names)
        return self._get_workbook_source(book)


if __name__ == '__main__':
    from Storage import Storage, AllowedDataTypes
    price_types = [item.value for item in AllowedDataTypes]
    exporter = ExportExcel(Storage(), price_types)
    source = exporter.export_to_excel()
    with open("sample.xlsx", 'wb') as fp:
        fp.write(source)
